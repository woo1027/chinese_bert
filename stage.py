import pymssql
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


# 建立連線
conn = pymssql.connect(
    server='192.168.0.104',
    user='sa',
    password='wu469711',
    database='student'
)



# 撰寫 SQL 查詢語句
sql = """
WITH login_count AS (
    SELECT student_id, COUNT(*) AS login_count
    FROM login_log
    GROUP BY student_id   /*登入次數*/
),
homework_stats AS (
    SELECT student_id,
           COUNT(*) AS total_hw,
           SUM(CAST(submitted AS INT)) AS submitted_hw
    FROM homework_submission 
    GROUP BY student_id   /*作業完成率*/
),
exam_stats AS (
    SELECT student_id, AVG(score) AS avg_score
    FROM exam_scores
    GROUP BY student_id    /*考試平均成績*/
)

SELECT s.student_id, s.name,
       ISNULL(l.login_count, 0) AS login_count,
       ISNULL(h.submitted_hw * 1.0 / NULLIF(h.total_hw, 0), 0) AS homework_completion_rate,
       ISNULL(e.avg_score, 0) AS average_score
FROM student s
LEFT JOIN login_count l ON s.student_id = l.student_id
LEFT JOIN homework_stats h ON s.student_id = h.student_id
LEFT JOIN exam_stats e ON s.student_id = e.student_id
-- 整合以上三張彙整表，加上學生基本資料

"""

df = pd.read_sql(sql, conn)
conn.close()



wb = Workbook()
ws = wb.active
ws.title = "Student 學習狀況表"

# 寫入資料
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 加入條件格式：成績<60填紅色
for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        if cell.value < 60:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# 儲存
wb.save("student_weekly_report.xlsx")

